# apps/students/services/import_xlsx.py
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from datetime import date, datetime

from django.db import transaction
from django.utils.dateparse import parse_date
from openpyxl import load_workbook

from apps.students.serializers import StudentWriteSerializer
from apps.students.models import Student


# --- Import contract (columns) -------------------------------------------------
# Required: first_name, last_name, date_of_birth
# Optional: gender, marital_status, phone_number, email, nationality, national_id,
# previous_institute, grade_acquired, district, county, sub_county_division,
# parish, cell_village, entry_date, exit_date, comments
#
# Optional (status bootstrap in future): status, term_name  -> can be ignored now.

REQUIRED = {"first_name", "last_name", "date_of_birth"}

CANONICAL_COLUMNS = [
    "first_name",
    "last_name",
    "date_of_birth",
    "gender",
    "marital_status",
    "phone_number",
    "email",
    "nationality",
    "national_id",
    "previous_institute",
    "grade_acquired",
    "district",
    "county",
    "sub_county_division",
    "parish",
    "cell_village",
    "entry_date",
    "exit_date",
    "comments",
]


@dataclass
class RowOutcome:
    row_number: int
    action: str  # "validated" | "created" | "skipped" | "error"
    errors: Dict[str, Any] = field(default_factory=dict)
    instance_id: Optional[int] = None


def _coerce_date(val: Any) -> Optional[date]:
    """
    Accepts Excel-native datetime/date or ISO-like strings. Returns date or None.
    """
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        # If someone drops a serial here, openpyxl usually already converts it;
        # we avoid guessing. Let serializer reject bad data.
        return None
    if isinstance(val, str):
        val = val.strip()
        # Normalize common dd/mm/yyyy or mm/dd/yyyy into ISO when unambiguous
        # Prefer Django’s parse_date (expects ISO). If not ISO, return None and let serializer handle.
        parsed = parse_date(val)
        return parsed
    return None


def _row_to_payload(headers: List[str], row_values: List[Any]) -> Dict[str, Any]:
    """
    Map a worksheet row into the StudentWriteSerializer payload.
    Unknown columns are ignored; missing required checked later.
    """
    data: Dict[str, Any] = {}
    for col_name, value in zip(headers, row_values):
        k = (col_name or "").strip().lower()
        if not k:
            continue
        if k not in set(CANONICAL_COLUMNS).union(REQUIRED):
            # ignore unknown columns silently (extensible)
            continue
        if k in {"date_of_birth", "entry_date", "exit_date"}:
            data[k] = _coerce_date(value)
        else:
            data[k] = value if value not in (None, "") else None
    return data


def _validate_required(payload: Dict[str, Any]) -> Dict[str, str]:
    errors: Dict[str, str] = {}
    for name in REQUIRED:
        if not payload.get(name):
            errors[name] = "This field is required."
    return errors


def import_students_xlsx(
    request,
    file_obj,
    *,
    commit: bool = False,
    atomic: bool = False,
) -> Dict[str, Any]:
    """
    Perform a dry-run (commit=False) or actual import (commit=True).
    If atomic=True and commit=True -> single transaction over all rows.
    Otherwise we use per-row transactions so good rows are not blocked by bad ones.

    Returns a summary with row-level outcomes.
    """
    iid = getattr(request.user, "institute_id", None)
    if not iid:
        return {
            "summary": {"created": 0, "validated": 0, "skipped": 0, "errors": 1},
            "rows": [
                RowOutcome(
                    row_number=0,
                    action="error",
                    errors={"institute": "User has no institute assigned."},
                ).__dict__
            ],
        }

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    # header row
    headers: List[str] = []
    first = True
    outcomes: List[RowOutcome] = []
    created = validated = skipped = errors = 0

    # Pre-create a consistent serializer context (respects your domain rules)
    context = {"request": request}

    def _process_row(row_idx: int, values: List[Any]) -> RowOutcome:
        payload = _row_to_payload(headers, values)
        req_errs = _validate_required(payload)
        if req_errs:
            return RowOutcome(row_number=row_idx, action="error", errors=req_errs)

        # Serializer will:
        # - enforce required fields again
        # - check dedup via has_potential_duplicate(...)
        # - set institute_id + spin
        # - create initial StudentStatus (ENQUIRE) with nearest term
        ser = StudentWriteSerializer(data=payload, context=context)
        if not ser.is_valid():
            return RowOutcome(row_number=row_idx, action="error", errors=ser.errors)

        if not commit:
            return RowOutcome(row_number=row_idx, action="validated")

        # Commit path
        try:
            obj: Student = ser.save()
            return RowOutcome(row_number=row_idx, action="created", instance_id=obj.id)
        except Exception as e:  # keep granular later if desired
            return RowOutcome(
                row_number=row_idx,
                action="error",
                errors={"non_field_errors": [str(e)]},
            )

    # Iterate rows
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if first:
            headers = [str(c or "").strip() for c in row]
            first = False
            # quick sanity check for required columns present
            missing = REQUIRED.difference({h.lower() for h in headers})
            if missing:
                msg = f"Missing required columns: {', '.join(sorted(missing))}"
                return {
                    "summary": {
                        "created": 0,
                        "validated": 0,
                        "skipped": 0,
                        "errors": 1,
                    },
                    "rows": [
                        RowOutcome(
                            row_number=1, action="error", errors={"header": msg}
                        ).__dict__
                    ],
                }
            continue

        # skip empty lines
        if not any(x not in (None, "", 0) for x in row):
            continue

        if commit and atomic:
            # single transaction around entire import (handled outside)
            outcome = _process_row(i, list(row))
        elif commit:
            # per-row transaction (good rows survive)
            with transaction.atomic():
                outcome = _process_row(i, list(row))
        else:
            # dry-run
            outcome = _process_row(i, list(row))

        outcomes.append(outcome)
        if outcome.action == "created":
            created += 1
        elif outcome.action == "validated":
            validated += 1
        elif outcome.action == "skipped":
            skipped += 1
        else:
            errors += 1

    if commit and atomic:
        # If any error and atomic file-level, rollback everything
        if errors > 0:
            raise transaction.TransactionManagementError(
                "Atomic import failed; nothing was created."
            )

    return {
        "summary": {
            "created": created,
            "validated": validated,
            "skipped": skipped,
            "errors": errors,
            "total_rows": len(outcomes),
            "commit": bool(commit),
            "atomic": bool(atomic),
        },
        "rows": [o.__dict__ for o in outcomes],
        "expected_columns": CANONICAL_COLUMNS,
    }
